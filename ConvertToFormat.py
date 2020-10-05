#!/usr/bin/env python

# export data sheets from xlsx to csv
from openpyxl import load_workbook
import csv, datetime, os, time
from os import sys

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file,read_only=True,data_only=True)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file,data_only=True)
    for worksheet_name in sheets:
        print("Экспорт " + worksheet_name + " ...")

        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
        except KeyError:
            print("Не возможно найти " + worksheet_name)
            sys.exit(1)
            
        your_csv_file = open(''.join([excel_file[:len(excel_file)-5],'.csv']), 'w', encoding='utf-8')
        wr = csv.writer(your_csv_file, delimiter='\t', lineterminator='\n')
        for row in worksheet.iter_rows(values_only=True):
            lrow = []
            for cell in row:
                if isinstance(cell, datetime.date)==True:
                    lrow.append(cell.strftime('%d.%m.%Y'))
                else:
                    lrow.append(cell)
            wr.writerow(lrow)
        print(" ... Готово!!!!")
        workbook.close()
        your_csv_file.close()


if not 2 <= len(sys.argv) <= 3:
    print("Call with " + sys.argv[0] + " <xlxs file> [comma separated list of sheets to export]")
    sys.exit(1)
else:
    sheets = []
    if len(sys.argv) == 3:
        sheets = list(sys.argv[2].split(','))
    else:
        sheets = get_all_sheets(sys.argv[1])
    assert(sheets != None and len(sheets) > 0)
    csv_from_excel(sys.argv[1], sheets)
